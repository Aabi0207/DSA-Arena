import openpyxl
import csv

def extract_from_excel(excel_path, output_tsv):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Find column indexes
    headers = [cell.value for cell in ws[1]]
    type_idx = headers.index("Type") + 1
    difficulty_idx = headers.index("Difficulty") + 1
    name_idx = headers.index("Name") + 1

    with open(output_tsv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(["Problem Name", "Problem Link", "Topic", "Solution Link", "Difficulty"])

        for row in ws.iter_rows(min_row=2):  # Skip header
            topic = row[type_idx - 1].value or ""
            difficulty = row[difficulty_idx - 1].value or ""

            # Problem Name and Link from the hyperlinked cell
            name_cell = row[name_idx - 1]
            problem_name = name_cell.value or ""
            problem_link = ""
            if name_cell.hyperlink:
                problem_link = name_cell.hyperlink.target

            writer.writerow([problem_name, problem_link, topic, problem_link, difficulty])

    print(f"Successfully written to {output_tsv}")

# Example usage
extract_from_excel(r'C:\Users\abhis\OneDrive\Desktop\Coding\FullStack\CapstoneProjects\DSA-Arena\DataCollector\blind75.xlsx', 'output.tsv')
