import openpyxl
import csv
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


# Load workbook and get all sheet names
wb = openpyxl.load_workbook(r"C:\Users\abhis\OneDrive\Desktop\Coding\FullStack\CapstoneProjects\DSA-Arena\DataCollector\topLeetcode.xlsx", data_only=True)

# Create a tab-separated output file
with open("mostAskedFAANG.tsv", "w", newline="", encoding="utf-8") as outfile:
    tsv_writer = csv.writer(outfile, delimiter="\t")
    
    # Write header
    tsv_writer.writerow(["Problem Name", "Problem Link", "Topic", "Solution Link", "Difficulty"])
    
    # Loop through each sheet
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        
        for row in sheet.iter_rows(min_row=6, max_row=135):
            # Get problem name and link from column B
            problem_cell = row[1]  # Column B
            problem_name = problem_cell.value if problem_cell.value else ""
            problem_link = problem_cell.hyperlink.target if problem_cell.hyperlink else ""
            
            # Get topic name from column C
            topic = row[2].value if row[2].value else ""
            
            # Get solution link from column D
            solution_link = row[3].hyperlink.target if row[3].hyperlink else (row[3].value or "")
            
            # Get difficulty from column J
            difficulty = row[9].value if row[9].value else ""
            
            # Write to TSV
            tsv_writer.writerow([problem_name, problem_link, topic, solution_link, difficulty])

print("✅ Extraction complete. Output saved as 'mostAskedFAANG.tsv'")
